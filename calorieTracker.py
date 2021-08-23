import myfitnesspal
from myfitnesspal.keyring_utils import get_password_from_keyring
from datetime import date
import openpyxl

#getting the current date
today = date.today()
def main():
    dailyWeight = getWeight()
    print (dailyWeight)

    dailyCal = calcCal()
    print (dailyCal)

    dailyShift()
    fileWrite(dailyWeight, dailyCal)

def getWeight():
    username = 'dan2001tran@gmail.com'
    client = myfitnesspal.Client(username, get_password_from_keyring(username))

    today = date.today()
    weight = client.get_measurements('Weight', today)

    return weight[today]
#logging into my fitnesspal and calculating the daily total calories from macros
def calcCal():
    username = 'dan2001tran@gmail.com'
    client = myfitnesspal.Client(username, get_password_from_keyring(username))

    day = client.get_date(today.year, today.month, today.day)
    

    calculatedDailyTotalCalories = day.totals['carbohydrates'] * 4 + day.totals['protein'] * 4 + day.totals['fat'] * 9
    return calculatedDailyTotalCalories

#will shift all past entries down one row
def dailyShift():

    #creating the workbook and worksheet object
    wb = openpyxl.load_workbook('/Users/dantran/Documents/Fitness/FitnessMeasurements.xlsx')
    ws = wb.active
    
    #formats the range of values that is going to be shifted and shifts
    targetCells = 'A4:C' + str(len(ws['A']))
    ws.move_range(targetCells,rows=1)
    
    wb.save('/Users/dantran/Documents/Fitness/FitnessMeasurements.xlsx')

#writes the new daily weight and daily calories
def fileWrite(weight, calories):
    #creating the workbook and worksheet object
    wb = openpyxl.load_workbook('/Users/dantran/Documents/Fitness/FitnessMeasurements.xlsx')
    ws = wb.active
    
    today = date.today()

    ws['A4'] = today
    ws['B4'] = weight
    ws['C4'] = calories

    wb.save('/Users/dantran/Documents/Fitness/FitnessMeasurements.xlsx')
main()